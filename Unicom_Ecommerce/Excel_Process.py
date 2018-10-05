import xlrd
from openpyxl import load_workbook,worksheet,cell,Workbook

num_dict = {}


def read_excel():
    wb = xlrd.open_workbook(u'D:/Desktop/订单明细.xlsx')
    first_sht = wb.sheet_by_index(0)
    first_row = first_sht.row_values(0)
    item_name_index = get_index(first_row, '商品名称')
    reserve_No_index = get_index(first_row, '商城预占号码')
    date_index = get_index(first_row, '订单日期')
    write_wb = Workbook()
    write_wb.create_sheet('16-18', index=5)
    write_sht = write_wb['16-18']
    write_sht.append(first_row)
    for sht_name in wb.sheet_names():
        sht = wb.sheet_by_name(sht_name)
        # first_row = sht.row_values(0)
        row_gener = sht.get_rows()
        # item_name_index = get_index(first_row,'商品名称')
        # reserve_No_index = get_index(first_row,'商城预占号码')
        # date_index = get_index(first_row,'订单日期')
        next(row_gener)
        generat_map(item_name_index,reserve_No_index,date_index,row_gener)
    write_excel(write_sht)
    write_wb.save('D:/Desktop/订单明细_待副卡.xlsx')



def write_excel(sht):
    for key,value in num_dict.items():
        sht.append(value)


def generat_map(item_name_index,reserve_No_index,date_index,row_generator):
    """
    :param reserve_NO_index: 商城预占号码的 index
    :param data_index:  订单日期的 index
    :param row_generator: 一个标签页中的行生成器 
    :return: 返回一个 { 订购号码:[每行内容]} 的字典
    """
    for row in row_generator:
        item_name_value = row[item_name_index].value
        reserve_No_value = row[reserve_No_index].value
        date_value = row[date_index].value
        # dict_key = item_name_value + reserve_No_value
        # ignor_list = ['王卡亲情卡',]

        """
        如果新的一行数据的日期比字典里的相同订购号码的日期大，则更新。否则添加到字典里。  
        """
        # print(type(date_value),type(num_dict.get(reserve_No_value)[date_index]))
        # print(date_value,num_dict.get(reserve_No_value)[date_index])
        # print(date_value>num_dict.get(reserve_No_value)[date_index])
        # print('---------------------')
        # if num_dict.get(reserve_No_value, None):                          # 判断字典里是否有商城预占号码的key
            # print(1)
            # print(float(date_value))

            # if float(date_value)>float(num_dict.get(reserve_No_value)[date_index]):         # 如果新的一行里的时间大于字典里已有的相同key的数据的时间，否则更新该key对应的数值
                # if num_dict.get(reserve_No_value)[item_name_index] != item_name_value:  # 如果新一行的套餐名称与字典里已有key对应数据的套餐名称不同，则将号码+套餐名称作为key，添加数据，否则删除该数据并更新
                #     num_dict[dict_key] = row
                # elif num_dict.get(dict_key,None) is not None:
                #     num_dict.pop(dict_key)
        num_dict[reserve_No_value] = [cell.value for cell in row]
        # else:
        #     # print(2)
        #     num_dict[reserve_No_value] = [cell.value for cell in row]


def get_index(first_row,title_name):
    return first_row.index(title_name)



read_excel()

