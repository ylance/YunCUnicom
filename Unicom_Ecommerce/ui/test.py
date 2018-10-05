from openpyxl import load_workbook
import re

def get_billing_no_date(base, no_picture_path):
    work_book = load_workbook('D:/Desktop/test.xlsx')
    sheet = work_book[work_book.sheetnames[0]]
    # content = ''
    billing_no_index = 0
    date_index = 0
    # count_index = 0
    for row in sheet.values:

        for cell in row:
            if '订单号' in cell:
                # print(row.index(cell))
                billing_no_index = row.index(cell)
            if '下单日期' in cell:
                # print(row.index(cell))
                date_index = row.index(cell)
            # cotent_list = re.findall(r'[^\"]+', cell)
            # if len(cotent_list) > 1:
            #     content = re.findall(r'[^\"]+', cell)[1]
            # if content == '订单号':
            #     # billing_no_index = row.index(cell)
            #     billing_no_index = 0
            # if content == '下单日期':
            #     # date_index = row.index(cell)
            #     date_index = 2
            #     continue
        # print(count_index, self.count)

        # if count_index != self.count:
        #     count_index += 1
        #     continue
        # count_index += 1
        # print(row)
        billing_no = re.findall(r'[^\"]+', row[billing_no_index])[1]
        date = re.findall(r'[^\"]+', row[date_index])[1][0:4] + '-' + re.findall(r'[^\"]+', row[date_index])[1][
                                                                      4:6] + '-' + \
               re.findall(r'[^\"]+', row[date_index])[1][6:8]
        print(date, billing_no)
        # if billing_no.isdigit():
        #     self.before_check(date, billing_no)
        #     self.start_check(base, date, no_picture_path)

get_billing_no_date('sdf','sdf')