from selenium import webdriver
from PIL import Image

from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import time
from face_contrast import face_contrast
import base64
import os
import winreg
from openpyxl import load_workbook
import re


def error_repeat(func):
    def inner(self, *args, **kwargs):
        try:

            func(self, *args, **kwargs)
        except Exception as e:
            # print(e)
            def alert_isexit():
                try:
                    alert_refresh = WebDriverWait(self.b, 3, 0.5).until(EC.alert_is_present())
                    return True, alert_refresh
                except TimeoutException as e:
                    return False, None
            flag_network, alert_network = alert_isexit()
            if flag_network:
                alert_network.accept()
            self.b.refresh()
            time.sleep(1)
            flag, alert_refreshed = alert_isexit()
            if flag:
                alert_refreshed.accept()
            self.count -= 1
            time.sleep(1)
            inner(self, *args, **kwargs)
    return inner


def get_desktop():
    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                         r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders')  # 利用系统的链表
    return winreg.QueryValueEx(key, "Desktop")[0]


def create_dir(create_path):
    if not os.path.exists(create_path):
        os.makedirs(create_path)
    return create_path


class AutoCheck:
    def __init__(self, username, password):
        self.username = username
        self.password = password
        # self.billing_date = billing_date
        # self.page_number = int(page_number)
        self.base_dir = create_dir(str(get_desktop())+'/自动抽检/')
        self.count = 1
        # self.url = 'https://www.baidu.com'
        self.url = 'http://admin.mall.10010.com/'
        self.stop = False
        self.b = webdriver.Chrome()
        self.no_check_list = ['王卡亲情卡', '商城冰激凌', '冰激凌团购', '冰激凌拼团', '冰激凌套餐【超值特惠版】','顺丰速运巴枪卡']
        emailaddress = '83533450@163.com'
        password = 'ningyun596123'
        pop3_server = 'pop.163.com'

    def open_web(self):
        self.b.maximize_window()
        self.b.get(self.url)

    def element_isexist(self, id):
        try:
            WebDriverWait(self.b, 60, 0.5).until(
                EC.presence_of_element_located((By.ID, id)))
            return True
        except Exception as e:
            # print(xpath + " is not found")
            return False

    def element_isexist_xpath(self, xpath):
        try:
            WebDriverWait(self.b, 60, 0.5).until(
                EC.presence_of_element_located((By.XPATH, xpath)))
            return True
        except Exception as e:
            # print(xpath + " is not found")
            return False

    def is_not_visible(self, locator, timeout=60):
        try:
            WebDriverWait(self.b, timeout, 0.5).until_not(EC.visibility_of_element_located((By.CLASS_NAME, locator)))
            return True
        except TimeoutException:
            return False

    def image_is_available(self, xpath):
        WebDriverWait(self.b, 60, 0.5).until(EC.presence_of_element_located((By.XPATH, xpath)))
        image = self.b.find_element_by_xpath(xpath)
        image_loaded = self.b.execute_script(
            "return arguments[0].complete && typeof arguments[0].naturalWidth != \"undefined\" && arguments[0].naturalWidth > 0",
            image)
        if not image_loaded:
            return False
        else:
            return True

    def wait_image_loaded(self, xpath, times, sleep_time):
        for i in range(times):
            if self.image_is_available(xpath):
                return True
            else:
                time.sleep(sleep_time)

    def screenshot_and_save(self, xpath, screen_path, save_path):
        self.b.save_screenshot(screen_path)
        if xpath == '//*[@id="main_img"]':
            location = self.b.find_element_by_xpath('//*[@id="main_img"]').location
            size = self.b.find_element_by_xpath('//*[@id="main_img"]').size
        else:
            location = self.b.find_element_by_xpath('//*[@id="fangdajing"]/div[2]/video').location
            size = self.b.find_element_by_xpath('//*[@id="fangdajing"]/div[2]/video').size
        left = location['x']
        top = location['y']
        right = location['x'] + size['width']
        bottom = location['y'] + size['height']
        a = Image.open(screen_path)
        im = a.crop((left, top, right, bottom))
        im.save(save_path)
        time.sleep(1)

    def get_path(self, basedir, filename):
        return os.path.join(basedir, filename)

    def element_isexist_xpath_shrot_time(self, xpath, how_long, how_often):
        try:
            WebDriverWait(self.b, how_long, how_often).until(
                EC.presence_of_element_located((By.XPATH, xpath)))
            return True
        except Exception as e:
            return False

    def close(self, ):
        close_btn = self.b.find_element_by_xpath('//*[@id="close-check"]')
        self.b.execute_script('arguments[0].click();', close_btn)

    def mark(self):
            WebDriverWait(self.b, 60, 0.5).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="real-name-verify"]/input[1]')))
            self.b.find_element_by_xpath('//*[@id="real-name-verify"]/input[1]').click()  # 点击‘是否符合实名制 是’radio
            WebDriverWait(self.b, 60, 0.5).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="phone-halt-verify"]/input[2]')))
            self.b.find_element_by_xpath('//*[@id="phone-halt-verify"]/input[2]').click()  # 点击‘是否停机  否’ radio
            WebDriverWait(self.b, 60, 0.5).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="verify-confirm"]')))
            self.b.find_element_by_xpath('//*[@id="verify-confirm"]').click()  # 点击 '确定‘  按钮
            # close_btn = self.b.find_element_by_xpath('//*[@id="close-check"]')
            # self.b.execute_script('arguments[0].click();', close_btn)

    def mark_unqualified(self):
        WebDriverWait(self.b, 60, 0.5).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="photo-verify"]/input[2]')))
        self.b.find_element_by_xpath('//*[@id="photo-verify"]/input[2]').click()
        WebDriverWait(self.b, 60, 0.5).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="real-name-verify"]/input[2]')))
        self.b.find_element_by_xpath('//*[@id="real-name-verify"]/input[2]').click()
        WebDriverWait(self.b, 60, 0.5).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="phone-halt-verify"]/input[1]')))
        self.b.find_element_by_xpath('//*[@id="phone-halt-verify"]/input[1]').click()
        WebDriverWait(self.b, 60, 0.5).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="contentDiv"]/div[3]/table[2]/tbody/tr[5]/td[2]/input[2]')))
        self.b.find_element_by_xpath('//*[@id="contentDiv"]/div[3]/table[2]/tbody/tr[5]/td[2]/input[2]').click()
        WebDriverWait(self.b, 60, 0.5).until(
            EC.presence_of_element_located(
                (By.XPATH, '//*[@id="verify-confirm"]')))
        self.b.find_element_by_xpath('//*[@id="verify-confirm"]').click()

    def mark_and_close(self, ):
        self.mark()
        self.close()

    def login(self):
        # 登陆开始
        WebDriverWait(self.b, 60, 0.5).until(EC.presence_of_element_located((By.ID, "merchantId")))
        self.b.find_element_by_id("merchantId").send_keys(self.username)
        WebDriverWait(self.b, 60, 0.5).until(EC.presence_of_element_located((By.ID, "merchantPwd")))
        self.b.find_element_by_id("merchantPwd").send_keys(self.password)

        # todo 短信验证码开始
        WebDriverWait(self.b, 60, 0.5).until(EC.presence_of_element_located((By.ID, "phone")))
        self.b.find_element_by_id('phone').click()
        WebDriverWait(self.b, 60, 0.5).until(EC.presence_of_element_located((By.ID, "merchantSafetyCodeBtn")))
        email_button = self.b.find_element_by_id("merchantSafetyCodeBtn")
        email_button.click()
        # todo 短信验证码结束

        # todo 邮箱验证码
        # WebDriverWait(b, 60, 0.5).until(EC.presence_of_element_located((By.ID, "email")))
        # b.find_element_by_id("email").click()
        # send_code_button = b.find_element_by_id('merchantSafetyCodeBtn')
        # # while b.title != '商城后台主页':
        # if send_code_button.get_attribute('class') == 'longBtn':
        #     send_code_button.click()
        #     time.sleep(30)
        #     code = get_email_code(emailaddress, password, pop3_server)
        #     b.find_element_by_id("merchantSafetyCode").clear()
        #     b.find_element_by_id("merchantSafetyCode").send_keys(code)
        #     WebDriverWait(b, 60, 0.5).until(EC.presence_of_element_located((By.ID, "merchantLogin")))
        #     b.find_element_by_id("merchantLogin").click()
        #     while element_isexist("merchantSafetyCodeCheck") or element_isexist('merchantSafetyCodeSendTips'):
        #         time.sleep(31)
        #         if b.title == '商城后台主页':
        #             break
        #         if send_code_button.get_attribute('class') == 'longBtn':
        #             send_code_button.click()
        #         code = get_email_code(emailaddress, password, pop3_server)
        #         b.find_element_by_id("merchantSafetyCode").clear()
        #         b.find_element_by_id("merchantSafetyCode").send_keys(code)
        #         WebDriverWait(b, 60, 0.5).until(EC.presence_of_element_located((By.ID, "merchantLogin")))
        #         b.find_element_by_id("merchantLogin").click()

    def select_menu(self):
        WebDriverWait(self.b, 180, 0.5).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="mainbBodyer"]/div[1]/div[9]/h3')))
        self.b.find_element_by_xpath('//*[@id="mainbBodyer"]/div[1]/div[9]/h3').click()
        WebDriverWait(self.b, 60, 0.5).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="mainbBodyer"]/div[1]/div[9]/h3')))
        select_button = self.b.find_element_by_xpath('//*[@id="mainbBodyer"]/div[1]/div[9]/h3')
        self.b.execute_script("arguments[0].click();", select_button)
        WebDriverWait(self.b, 60, 0.5).until(EC.presence_of_element_located((By.ID, "menu_5600")))
        distribute_link = self.b.find_element_by_id("menu_5600")
        self.b.execute_script("arguments[0].click();", distribute_link)

    def before_check(self, billing_date, orderNo):
        # todo 选择购买时间，可选项
        WebDriverWait(self.b, 60, 0.5).until(EC.presence_of_element_located((By.ID, "maxTime")))
        # time_input = b.find_element_by_id('maxTime')
        self.b.execute_script('document.getElementById("maxTime").removeAttribute("onfocus");')
        self.b.find_element_by_id('maxTime').clear()
        current_date = time.strftime('%Y-%m-%d', time.localtime(time.time()))
        if billing_date is not None or billing_date.strip() != '':
            self.b.find_element_by_id('maxTime').send_keys(billing_date)
        else:
            self.b.find_element_by_id('maxTime').send_keys(current_date)

        # todo 选择订单状态，将其选为成功关闭
        # WebDriverWait(self.b, 60, 0.5).until(EC.presence_of_element_located((By.ID, "orderState")))
        # state = self.b.find_element_by_id('orderState')
        # Select(state).select_by_value('00')
        # todo 输入单号
        WebDriverWait(self.b, 60, 0.5).until(EC.presence_of_element_located((By.ID, "orderNo")))
        self.b.find_element_by_id('orderNo').clear()
        self.b.find_element_by_id('orderNo').send_keys(orderNo)

        WebDriverWait(self.b, 60, 0.5).until(EC.element_to_be_clickable((By.ID, 'queryBtn')))
        query_btn = self.b.find_element_by_id('queryBtn')
        self.b.execute_script('arguments[0].click();', query_btn)
        self.is_not_visible("thickdiv")
        WebDriverWait(self.b, 60, 0.5).until_not(EC.visibility_of_element_located((By.XPATH, '//*[@id="pagination_thickdiv"]')))

    def get_billing_no_date(self, base, no_picture_path):
        work_book = load_workbook('C:/Users/Administrator/Desktop/test.xlsx')
        sheet = work_book[work_book.sheetnames[0]]
        content = ''
        billing_no_index = 0
        date_index = 0
        count_index = 0
        for row in sheet.values:

            for cell in row:
                cotent_list = re.findall(r'[^\"]+', cell)
                if len(cotent_list) > 1:
                    content = re.findall(r'[^\"]+', cell)[1]
                if content == '订单号':
                    # billing_no_index = row.index(cell)
                    billing_no_index =2
                if content == '下单日期':
                    # date_index = row.index(cell)
                    date_index = 4
                    continue
            # print(count_index, self.count)

            if count_index != self.count:
                count_index += 1
                continue
            count_index += 1
            # print(row)
            billing_no = re.findall(r'[^\"]+', row[billing_no_index])[1]
            date = re.findall(r'[^\"]+', row[date_index])[1][0:4] + '-' + re.findall(r'[^\"]+', row[date_index])[1][
                                                                          4:6] + '-' + \
                   re.findall(r'[^\"]+', row[date_index])[1][6:8]
            # print(date, billing_no)
            if billing_no.isdigit():

                self.before_check(date, billing_no)
                self.start_check(base, date,no_picture_path)

    def start_check(self, base, billing_date, no_picture_path):
        # todo 数字可变
        temp_path = self.get_path(create_dir(base + 'temp/'), 'temp.png')
        # todo 晓雯
        identity_base = create_dir(base + '抽检复查/第二次/' + billing_date + '/身份证/')
        user_base = create_dir(base + '抽检复查/第二次/' + billing_date + '/用户照片/')
        # identity_base = create_dir(base + billing_date + '/身份证/')
        # user_base = create_dir(base + billing_date + '/用户照片/')

        self.count += 1
        print('第'+str(self.count-1)+'个开始')
        # if self.count == 10:
        #     self.b.quit()
        #     print('尝试次数达到10次，请重新开始')

        self.b.execute_script('document.documentElement.scrollTop=10;')
        write_path = create_dir(self.base_dir + '抽检复查/第二次/')
        if self.element_isexist_xpath_shrot_time('//*[@id="pageTable"]/div[2]/table/tbody/tr/td[5]/div', 2, 0.2):
            # print('-----第'+str(self.page_number)+'页开始-----')
            table = self.b.find_element_by_xpath('//*[@id="pageTable"]')
            divs = table.find_elements_by_class_name('tableBody')
            page_size = len(divs)
            for i in range(page_size):
                billing_no = self.b.find_element_by_xpath('//*[@id="pageTable"]/div['+str(i+2)+']/table/tbody/tr/td[1]/p[2]').text.split('：')[-1]
                if self.b.find_element_by_xpath('//*[@id="pageTable"]/div['+str(i+2)+']/table/tbody/tr/td[2]/p').text in self.no_check_list:
                    text = self.b.find_element_by_xpath('//*[@id="pageTable"]/div['+str(i+2)+']/table/tbody/tr/td[2]/p').text
                    tencent_qinqing = open(write_path + '需人工核实单号.txt', 'a')
                    tencent_qinqing.write(billing_date+' '+ billing_no + ' '+ text +'\n')
                    print(billing_no+'为：'+text)
                    tencent_qinqing.close()
                    continue
                identity_path = identity_base + billing_no + '.jpg'
                user_path = user_base + billing_no + '.jpg'
                if self.element_isexist_xpath_shrot_time('//*[@id="pageTable"]/div['+str(i+2)+']/table/tbody/tr/td[1]/p[4]/span[2]', 1, 0.2) or \
                        self.element_isexist_xpath_shrot_time('//*[@id="pageTable"]/div[' + str(i + 2) + ']/table/tbody/tr/td[1]/p[5]/span[2]', 1, 0.2):
                    continue
                WebDriverWait(self.b, 60, 0.5).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="pageTable"]/div[' + str(i+2) + ']/table/tbody/tr/td[5]/div')))
                check_btn = self.b.find_element_by_xpath('//*[@id="pageTable"]/div[' + str(i+2) + ']/table/tbody/tr/td[5]/div')
                self.b.execute_script("arguments[0].click();", check_btn)  # 点击‘抽检’ 按钮

                if self.b.find_element_by_xpath('//*[@id="contentDiv"]/div[3]/table[1]').is_displayed():
                    self.close()
                    continue

                if self.wait_image_loaded('//*[@id="cert-img-front"]', 60, 0.5):
                    img = self.b.find_element_by_xpath('//*[@id="cert-img-front"]')
                    self.b.execute_script("arguments[0].click();", img)
                    self.screenshot_and_save('//*[@id="main_img"]',temp_path, identity_path)  # 保存第一个

                simeliar_text = self.b.find_element_by_xpath('//*[@id="verifySimilarity"]').text.split('：')[-1].split('%')[0]
                if self.b.find_element_by_xpath('//*[@id="verifySimilarity"]').is_displayed() and float(
                        simeliar_text) > 70.0:
                    print(billing_no + '活体相似度为：'+ str(simeliar_text))
                    self.mark_and_close()
                    continue

                if self.b.find_element_by_xpath('//*[@id="showArea"]/div[8]').is_displayed():
                    self.b.execute_script('document.getElementById("showArea").scrollTop = 1000;')
                    img = self.b.find_element_by_xpath('//*[@id="gztliving_1"]')
                    self.b.execute_script('arguments[0].click()',img)
                    self.screenshot_and_save('//*[@id="main_img"]', temp_path, user_path)
                    confidence, errno = face_contrast(identity_path, user_path)
                    print(billing_no+'相似度为：',confidence)
                    if confidence >= 20:
                        self.mark_and_close()
                    elif confidence < 20:
                        self.mark_unqualified()
                        non_standard = open(write_path + '需人工核实单号.txt', 'a')
                        non_standard.write(billing_date+' '+ billing_no +' '+str(confidence)+'\n')
                        non_standard.close()
                        self.close()
                elif self.b.find_element_by_xpath('//*[@id="showArea"]/div[7]').is_displayed():
                    self.b.execute_script('document.getElementById("showArea").scrollTop = 1000;')
                    img = self.b.find_element_by_xpath('//*[@id="living-photo"]')  # 6
                    self.b.execute_script('arguments[0].click()', img)
                    self.screenshot_and_save('//*[@id="main_img"]',temp_path, user_path)
                    with open(user_path, 'rb') as user_ph, open(identity_path,'rb') as ident_ph, open(no_picture_path, 'rb') as no_picture_ph:
                        if base64.b64encode(user_ph.read()) != base64.b64encode(no_picture_ph.read()):
                            confidence, errno = face_contrast(identity_path, user_path)
                            print(billing_no+'相似度为：', confidence)
                            if confidence >= 20:
                                self.mark_and_close()
                            elif confidence < 20:
                                self.mark_unqualified()
                                non_standard = open(write_path + '需人工核实单号.txt', 'a')
                                non_standard.write('视频截图不合格: '+billing_date+' '+ billing_no +' '+str(confidence)+'\n')
                                non_standard.close()
                                self.close()
                        else:
                            img = self.b.find_element_by_xpath('//*[@id="living-video"]')  # 7
                            self.b.execute_script('arguments[0].click()', img)
                            time.sleep(1.5)
                            self.screenshot_and_save('//*[@id="fangdajing"]/div[2]/video',temp_path, user_path)
                            with open(user_path, 'rb') as user_ph:
                                confidence, errno = face_contrast(identity_path, user_path)
                                print(billing_no+'相似度为：', confidence)
                                if confidence >= 20:
                                    self.mark_and_close()
                                elif confidence < 20:
                                    self.mark_unqualified()
                                    non_standard = open(write_path + '需人工核实单号.txt', 'a')
                                    non_standard.write('视频不合格: '+billing_date+' '+ billing_no +' '+str(confidence)+'\n')
                                    non_standard.close()
                                    self.close()
                else:
                    img = self.b.find_element_by_id("cert-card-img")
                    self.b.execute_script("arguments[0].click();", img)
                    self.screenshot_and_save('//*[@id="main_img"]', temp_path, user_path)  # 保存第四个
                    confidence, errno = face_contrast(identity_path, user_path)
                    print(billing_no+'相似度为：', confidence)
                    if confidence >= 20:
                        self.mark_and_close()
                    elif confidence < 20:
                        self.mark_unqualified()
                        non_standard = open(write_path + '需人工核实单号.txt', 'a')
                        non_standard.write('照片: '+billing_date+' '+ billing_no +' '+str(confidence)+'\n')
                        non_standard.close()
                        self.close()

    def start(self):
        self.open_web()
        self.login()
        self.select_menu()

    def quit_app(self):
        self.b.quit()
        raise Exception

    @error_repeat
    def main_pro(self):
        self.get_billing_no_date(self.base_dir,self.get_path(self.base_dir+'temp/', 'no_picture.jpg'))
        # self.before_check(self.billing_date)
        # self.start_check(self.base_dir, self.billing_date, self.get_path(self.base_dir+'temp/', 'no_picture.jpg'))


if __name__ =='__main__':
    username = 'SXYC0032'
    password = 'jiahuan123+'
    auto = AutoCheck(username,password)
    auto.start()
    auto.main_pro()
    # def deamon(self):
    #     self.start()
    #     self.main_pro()
    #     self.b.quit()
    #     print('结束')
    #
    # def main(self):
    #     while True:
    #         if self.stop is True:
    #             self.b.quit()
    #             print('结束')
    #             break
